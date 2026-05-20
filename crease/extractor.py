"""Apply a Template to an xlsx file, produce canonical JSON."""

from __future__ import annotations

import re
from collections.abc import Iterator
from dataclasses import dataclass
from dataclasses import field as _dc_field
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from crease._coerce import (
    CoercionError,
    coerce,
    collapse_null,
    normalize_header,
    normalize_value,
    resolve_null_tokens,
)
from crease._locate import (
    CellRange,
    TabMatch,
    find_tabs,
    hidden_row_indices,
    normalize_headers,
    parse_cell_range,
    resolve_header_row,
)
from crease.template_model import (
    DataEnd,
    Enrich,
    Entity,
    Template,
)

# ---- result types --------------------------------------------------------


@dataclass
class ExtractionError:
    entity: str | None
    reason: str
    details: dict[str, Any] = _dc_field(default_factory=dict)


@dataclass
class RowExtractError:
    """A failure to coerce a single cell. Lifted into a validation Issue downstream."""

    entity: str
    row: int
    field: str
    reason: str  # "wrong_type" | "missing_required" | "anchor_not_found"
    expected: str | None = None
    got: Any = None
    likely_cause: str | None = None


@dataclass
class ExtractResult:
    template_id: str
    source_file: str
    canonical: dict[str, Any] = _dc_field(default_factory=dict)
    errors: list[ExtractionError] = _dc_field(default_factory=list)
    row_errors: list[RowExtractError] = _dc_field(default_factory=list)


# ---- helpers -------------------------------------------------------------


def _pluralize(name: str) -> str:
    if name.endswith("s"):
        return name
    if name.endswith("y") and not name.endswith(("ay", "ey", "oy", "uy")):
        return name[:-1] + "ies"
    return name + "s"


def _open_workbook(path: Path) -> Workbook:
    return openpyxl.load_workbook(path, data_only=True, read_only=False)


def _filename_inject(record: dict[str, Any], template: Template, source_file: str) -> dict[str, Any]:
    if not template.filename_pattern or not template.filename_capture:
        return record
    m = re.match(template.filename_pattern, source_file)
    if not m:
        return record
    for cap in template.filename_capture:
        try:
            raw = m.group(cap.group)
            record[cap.field] = raw  # type coercion deferred for now
        except IndexError:
            continue
    return record


def _apply_enrich(record: dict[str, Any], enrich_list: list[Enrich], tab: TabMatch) -> dict[str, Any]:
    for enrich in enrich_list:
        if enrich.source == "tab_name":
            raw = tab.name
        elif enrich.source == "tab_name_regex_group":
            if tab.regex_match is None:
                raw = None
            else:
                try:
                    raw = tab.regex_match.group(enrich.group)
                except IndexError:
                    raw = None
        else:
            raw = None
        if raw is not None and isinstance(raw, str):
            if enrich.strip_prefix:
                raw = raw.removeprefix(enrich.strip_prefix)
            if enrich.strip_suffix:
                raw = raw.removesuffix(enrich.strip_suffix)
        record[enrich.field] = raw
    return record


# ---- flat orientation ----------------------------------------------------


def _read_flat_grid(ws: Worksheet, locate, cell_range: CellRange | None) -> list[list[Any]]:
    """Pull rows into a 2D list. openpyxl read; pandas not used yet (kept simple)."""
    grid: list[list[Any]] = []
    hidden = hidden_row_indices(ws) if locate.skip_hidden_rows else set()

    min_row = (cell_range.start_row + 1) if cell_range else 1
    max_row = (cell_range.end_row + 1) if (cell_range and cell_range.end_row is not None) else None
    min_col = (cell_range.start_col + 1) if cell_range else 1
    max_col = (cell_range.end_col + 1) if (cell_range and cell_range.end_col is not None) else None

    iter_kwargs: dict[str, Any] = {"values_only": True, "min_row": min_row, "min_col": min_col}
    if max_row is not None:
        iter_kwargs["max_row"] = max_row
    if max_col is not None:
        iter_kwargs["max_col"] = max_col

    base = min_row - 1
    for offset, row in enumerate(ws.iter_rows(**iter_kwargs)):
        if (base + offset) in hidden:
            continue
        grid.append(list(row))

    while grid and all(v is None for v in grid[-1]):
        grid.pop()
    return grid


def _apply_data_end(rows: list[list[Any]], header_idx: int, data_end: DataEnd | None) -> list[list[Any]]:
    """Trim trailing rows based on data_ends_at."""
    if data_end is None or data_end.type == "end_of_sheet":
        return rows

    data = rows[header_idx + 1 :]

    if data_end.type == "blank_row":
        n = data_end.n_consecutive
        run = 0
        cutoff = len(data)
        for i, r in enumerate(data):
            if all(v is None for v in r):
                run += 1
                if run >= n:
                    cutoff = i - n + 1
                    break
            else:
                run = 0
        return rows[: header_idx + 1] + data[:cutoff]

    if data_end.type == "value_match":
        col = data_end.column
        value = data_end.value or ""
        for i, r in enumerate(data):
            if col < len(r) and r[col] is not None and value in str(r[col]):
                return rows[: header_idx + 1] + data[:i]
        return rows

    if data_end.type == "skip_trailing_rows":
        n = data_end.rows
        return rows[:-n] if n > 0 else rows

    return rows


def _extract_flat(
    ws: Worksheet,
    entity: Entity,
    tab: TabMatch,
    template: Template,
    result: ExtractResult,
) -> list[dict[str, Any]] | dict[str, Any] | None:
    cell_range = parse_cell_range(entity.locate.cell_range) if entity.locate.cell_range else None
    grid = _read_flat_grid(ws, entity.locate, cell_range)

    header_idx = resolve_header_row(ws, entity.locate)
    if cell_range:
        header_idx = max(0, header_idx - cell_range.start_row)

    if header_idx >= len(grid):
        result.errors.append(
            ExtractionError(
                entity=entity.name,
                reason="entity_missing",
                details={"hint": "header_row beyond available rows"},
            )
        )
        return None

    grid = _apply_data_end(grid, header_idx, entity.locate.data_ends_at)
    raw_headers = grid[header_idx] if grid else []
    headers = normalize_headers(raw_headers)

    data_starts_offset = (
        (entity.locate.data_starts_row - (cell_range.start_row if cell_range else 0))
        if entity.locate.data_starts_row is not None
        else header_idx + 1
    )
    data_rows = grid[data_starts_offset:]

    # Map field → header column index
    field_to_col: dict[str, int] = {}
    missing_columns: list[str] = []
    for f in entity.fields:
        if f.source_column is None:
            continue
        wanted = normalize_header(f.source_column)
        try:
            field_to_col[f.name] = headers.index(wanted)
        except ValueError:
            missing_columns.append(f.source_column)

    if missing_columns:
        result.errors.append(
            ExtractionError(
                entity=entity.name,
                reason="header_mapping_failed",
                details={"missing": missing_columns, "got": list(raw_headers)},
            )
        )

    extracted: list[dict[str, Any]] = []
    null_token_cache = {f.name: resolve_null_tokens(f, template) for f in entity.fields}

    for r_idx, row in enumerate(data_rows):
        if all(v is None for v in row):
            continue  # skip fully blank rows mid-data; validator may flag

        record: dict[str, Any] = {}
        for f in entity.fields:
            col = field_to_col.get(f.name)
            raw = row[col] if (col is not None and col < len(row)) else None
            value = collapse_null(raw, null_token_cache[f.name])
            value = normalize_value(value, f.normalize)

            if value is None:
                record[f.name] = None
                if not f.nullable and col is not None:
                    result.row_errors.append(
                        RowExtractError(
                            entity=entity.name,
                            row=r_idx,
                            field=f.name,
                            reason="missing_required",
                            expected=f.type,
                            got=None,
                        )
                    )
                continue

            try:
                record[f.name] = coerce(value, f)
            except CoercionError as e:
                record[f.name] = value
                result.row_errors.append(
                    RowExtractError(
                        entity=entity.name,
                        row=r_idx,
                        field=f.name,
                        reason="wrong_type",
                        expected=e.expected,
                        got=repr(value),
                        likely_cause=e.likely_cause,
                    )
                )

        record = _apply_enrich(record, entity.enrich, tab)
        record = _filename_inject(record, template, result.source_file)
        extracted.append(record)

    return extracted


# ---- property_sheet orientation ------------------------------------------


def _extract_property_sheet(
    ws: Worksheet,
    entity: Entity,
    tab: TabMatch,
    template: Template,
    result: ExtractResult,
) -> dict[str, Any]:
    """Read [label, value] pairs from two adjacent columns."""
    locate = entity.locate
    cell_range = parse_cell_range(locate.cell_range) if locate.cell_range else None

    label_col = locate.label_col
    value_col = locate.value_col
    start_row = locate.start_row

    if cell_range:
        label_col = cell_range.start_col
        value_col = cell_range.start_col + 1
        start_row = cell_range.start_row

    label_to_value: dict[str, Any] = {}
    end_row = (cell_range.end_row + 1) if (cell_range and cell_range.end_row is not None) else None

    iter_kwargs: dict[str, Any] = {"values_only": True, "min_row": start_row + 1}
    if end_row is not None:
        iter_kwargs["max_row"] = end_row

    for row in ws.iter_rows(**iter_kwargs):
        if label_col >= len(row):
            continue
        label = row[label_col]
        if label is None:
            continue
        value = row[value_col] if value_col < len(row) else None
        label_to_value[normalize_header(label).rstrip(":")] = value

    record: dict[str, Any] = {}
    for f in entity.fields:
        if f.source_label is None:
            continue
        key = normalize_header(f.source_label).rstrip(":")
        raw = label_to_value.get(key)
        nulls = resolve_null_tokens(f, template)
        raw = collapse_null(raw, nulls)
        raw = normalize_value(raw, f.normalize)
        if raw is None:
            record[f.name] = None
            if not f.nullable:
                result.row_errors.append(
                    RowExtractError(
                        entity=entity.name,
                        row=0,
                        field=f.name,
                        reason="missing_required",
                        expected=f.type,
                        got=None,
                    )
                )
            continue
        try:
            record[f.name] = coerce(raw, f)
        except CoercionError as e:
            record[f.name] = raw
            result.row_errors.append(
                RowExtractError(
                    entity=entity.name,
                    row=0,
                    field=f.name,
                    reason="wrong_type",
                    expected=e.expected,
                    got=repr(raw),
                    likely_cause=e.likely_cause,
                )
            )

    record = _apply_enrich(record, entity.enrich, tab)
    record = _filename_inject(record, template, result.source_file)
    return record


# ---- anchored orientation ------------------------------------------------


def _extract_anchored(
    ws: Worksheet,
    entity: Entity,
    tab: TabMatch,
    template: Template,
    result: ExtractResult,
) -> dict[str, Any]:
    """Locate each field independently via its `anchor` spec."""
    grid: list[list[Any]] = []
    for row in ws.iter_rows(values_only=True):
        grid.append(list(row))

    record: dict[str, Any] = {}
    for f in entity.fields:
        if f.anchor is None:
            continue
        loc = _find_anchor(grid, f.anchor)
        if loc is None:
            record[f.name] = None
            result.row_errors.append(
                RowExtractError(
                    entity=entity.name,
                    row=0,
                    field=f.name,
                    reason="anchor_not_found",
                    expected=f.anchor.label_match,
                )
            )
            continue
        r, c = loc
        raw = _walk_from(grid, r, c, f.anchor)
        nulls = resolve_null_tokens(f, template)
        raw = collapse_null(raw, nulls)
        raw = normalize_value(raw, f.normalize)
        if raw is None:
            record[f.name] = None
            if not f.nullable:
                result.row_errors.append(
                    RowExtractError(
                        entity=entity.name,
                        row=0,
                        field=f.name,
                        reason="missing_required",
                        expected=f.type,
                    )
                )
            continue
        try:
            record[f.name] = coerce(raw, f)
        except CoercionError as e:
            record[f.name] = raw
            result.row_errors.append(
                RowExtractError(
                    entity=entity.name,
                    row=0,
                    field=f.name,
                    reason="wrong_type",
                    expected=e.expected,
                    got=repr(raw),
                    likely_cause=e.likely_cause,
                )
            )

    record = _apply_enrich(record, entity.enrich, tab)
    record = _filename_inject(record, template, result.source_file)
    return record


def _find_anchor(grid: list[list[Any]], anchor) -> tuple[int, int] | None:
    target = anchor.label_match
    mode = anchor.match_mode
    for r, row in enumerate(grid):
        for c, val in enumerate(row):
            if val is None:
                continue
            s = str(val).strip()
            if mode == "exact" and s == target:
                return (r, c)
            if mode == "contains" and target in s:
                return (r, c)
            if mode == "regex" and re.search(target, s):
                return (r, c)
    return None


def _walk_from(grid: list[list[Any]], r: int, c: int, anchor) -> Any:
    direction = anchor.value_at
    offset = anchor.offset
    if direction == "right":
        c += offset
    elif direction == "left":
        c -= offset
    elif direction == "below":
        r += offset
    elif direction == "above":
        r -= offset
    if 0 <= r < len(grid) and 0 <= c < len(grid[r]):
        return grid[r][c]
    return None


# ---- unpivot -------------------------------------------------------------


def _apply_unpivot(records: list[dict[str, Any]], entity: Entity) -> list[dict[str, Any]]:
    if entity.unpivot is None:
        return records
    up = entity.unpivot
    id_cols = set(up.id_columns)
    var_re = re.compile(up.variable_column_pattern)

    out: list[dict[str, Any]] = []
    for rec in records:
        for key, val in rec.items():
            if key in id_cols:
                continue
            if not var_re.match(str(key)):
                continue
            row = {k: rec[k] for k in id_cols if k in rec}
            row[up.variable_name] = key
            row[up.value_name] = val
            out.append(row)
    return out


# ---- entity dispatch -----------------------------------------------------


def _extract_entity(workbook: Workbook, entity: Entity, template: Template, result: ExtractResult) -> None:
    tabs = find_tabs(workbook, entity.locate, template.ignore_tabs)
    if not tabs:
        result.errors.append(
            ExtractionError(
                entity=entity.name,
                reason="tab_pattern_no_match" if entity.locate.tab_pattern else "missing_tab",
                details={
                    "tab": entity.locate.tab,
                    "tab_pattern": entity.locate.tab_pattern,
                },
            )
        )
        return

    orientation = entity.locate.orientation
    key = entity.name if entity.cardinality == "one" else _pluralize(entity.name)

    if entity.cardinality == "one":
        ws = tabs[0].worksheet
        tab = tabs[0]
        if orientation == "property_sheet":
            record = _extract_property_sheet(ws, entity, tab, template, result)
        elif orientation == "anchored":
            record = _extract_anchored(ws, entity, tab, template, result)
        elif orientation == "flat":
            rows = _extract_flat(ws, entity, tab, template, result)
            if isinstance(rows, list) and rows:
                if len(rows) > 1:
                    result.errors.append(
                        ExtractionError(
                            entity=entity.name,
                            reason="multiple_rows_for_cardinality_one",
                            details={"n_rows": len(rows)},
                        )
                    )
                record = rows[0]
            else:
                record = None
        else:
            result.errors.append(
                ExtractionError(
                    entity=entity.name,
                    reason="unsupported_orientation",
                    details={"orientation": orientation},
                )
            )
            return
        result.canonical[key] = record
        return

    # cardinality == "many" — aggregate across matching tabs
    all_rows: list[dict[str, Any]] = []
    for tab in tabs:
        ws = tab.worksheet
        if orientation == "flat":
            rows = _extract_flat(ws, entity, tab, template, result) or []
            all_rows.extend(rows)
        elif orientation == "property_sheet":
            all_rows.append(_extract_property_sheet(ws, entity, tab, template, result))
        elif orientation == "anchored":
            all_rows.append(_extract_anchored(ws, entity, tab, template, result))
        else:
            result.errors.append(
                ExtractionError(
                    entity=entity.name,
                    reason="unsupported_orientation",
                    details={"orientation": orientation},
                )
            )

    all_rows = _apply_unpivot(all_rows, entity)
    result.canonical[key] = all_rows


# ---- public API ----------------------------------------------------------


def extract(path: str | Path, template: Template) -> ExtractResult:
    """Apply a template to a file. Returns canonical JSON + structural errors."""
    p = Path(path)
    result = ExtractResult(template_id=template.template_id, source_file=p.name)
    workbook = _open_workbook(p)
    try:
        for entity in template.entities:
            _extract_entity(workbook, entity, template, result)
    finally:
        workbook.close()
    return result


def get(path: str | Path, template: Template, entity: str) -> Any:
    """Extract a single entity. Returns the canonical value for that entity."""
    result = extract(path, template)
    if entity in result.canonical:
        return result.canonical[entity]
    plural = _pluralize(entity)
    return result.canonical.get(plural)


def stream(path: str | Path, template: Template, entity: str) -> Iterator[dict[str, Any]]:
    """Yield records of one entity. For cardinality=many, true streaming over tab(s)."""
    # For v1, delegate to extract() and iterate. True row-by-row streaming
    # via openpyxl read_only is a follow-on once the eager path is stable.
    value = get(path, template, entity)
    if value is None:
        return
    if isinstance(value, list):
        yield from value
    else:
        yield value
