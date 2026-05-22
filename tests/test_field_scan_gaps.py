"""Failing tests for design gaps surfaced in the 2026-05-21 field scan.

Each test exercises one gap with a minimal, PII-free fixture and the
*proposed* YAML syntax (or proposed API). Tests are marked
``xfail(strict=True)`` so the suite stays green until each gap ships;
removing an xfail marker graduates the test to a real acceptance check.

Source: per-customer field-scan reports (out-of-tree). Gap IDs (P0-1,
P0-2, ...) match the synthesis doc. P0-1 is intentionally not retested
here because ``test_cases/repeating_sections_per_tab/`` already encodes
that proposed ``blocks:`` syntax as a fixture.

All data is fictitious — Acme/Globex/Hooli/Initech, ``example.com``,
``ORD-####``, generic farm/site names. No real customer/vendor data
crosses into this repo (see CLAUDE.md "No real PII in generated
artifacts").
"""

from __future__ import annotations

import textwrap
import zipfile
from collections.abc import Callable
from datetime import date, time
from pathlib import Path

import pytest
from openpyxl import Workbook

import crease
from crease import Template, extract, validate

# ----------------------------------------------------------------------
# helpers
# ----------------------------------------------------------------------


def _xlsx(
    tmp_path: Path,
    build: Callable[[Workbook], None],
    *,
    name: str = "input.xlsx",
) -> Path:
    """Build a workbook with the caller's setup, save it, return the path.

    The default sheet is removed so the caller controls sheet names and
    order — most tests want a single named tab.
    """
    wb = Workbook()
    default = wb.active
    wb.remove(default)
    build(wb)
    path = tmp_path / name
    wb.save(path)
    return path


def _yml(tmp_path: Path, body: str, *, name: str = "template.yml") -> Path:
    path = tmp_path / name
    path.write_text(textwrap.dedent(body).strip() + "\n")
    return path


def _load(tmp_path: Path, yml_body: str) -> Template:
    return Template.load(_yml(tmp_path, yml_body))


def _run(xlsx: Path, yml_body: str, tmp_path: Path):
    tmpl = _load(tmp_path, yml_body)
    return extract(xlsx, tmpl)


# ======================================================================
# P0-2  skip_row_if predicate (CONVENTIONS §6; deferred to v1.5)
# ======================================================================


def test_skip_row_if_drops_subtotal_rows(tmp_path):
    """Subtotal rows with a blank discriminator column should be filtered out."""

    def build(wb):
        ws = wb.create_sheet("Orders")
        ws.append(["customer", "qty", "amount"])
        ws.append(["Acme Co.", 5, 100.00])
        ws.append(["Globex Corp", 3, 60.00])
        ws.append([None, None, 160.00])  # subtotal: blank `customer`
        ws.append(["Hooli Inc.", 7, 140.00])

    xlsx = _xlsx(tmp_path, build)
    result = _run(
        xlsx,
        """
        template_id: skip_row_if_subtotals
        version: 1
        description: drop subtotal rows missing the discriminator column
        entities:
          - name: order
            cardinality: many
            locate:
              tab: Orders
              orientation: flat
              header_row: 0
              skip_row_if:
                - { all_blank: [customer] }
            fields:
              - { name: customer, source_column: customer, type: string }
              - { name: qty, source_column: qty, type: integer }
              - { name: amount, source_column: amount, type: number }
        """,
        tmp_path,
    )

    assert len(result.canonical["orders"]) == 3
    assert {r["customer"] for r in result.canonical["orders"]} == {
        "Acme Co.",
        "Globex Corp",
        "Hooli Inc.",
    }


def test_skip_row_if_drops_day_of_week_marker_rows(tmp_path):
    """Day-of-week marker rows (col 0 = MONDAY|TUESDAY|...) should be dropped."""

    def build(wb):
        ws = wb.create_sheet("Schedule")
        ws.append(["label", "qty"])
        ws.append(["MONDAY", None])  # marker, not data
        ws.append(["Acme Co.", 10])
        ws.append(["Globex Corp", 20])
        ws.append(["TUESDAY", None])  # marker, not data
        ws.append(["Hooli Inc.", 30])

    xlsx = _xlsx(tmp_path, build)
    result = _run(
        xlsx,
        """
        template_id: skip_row_if_day_markers
        version: 1
        description: P0-2 fixture - drop day-of-week marker rows
        entities:
          - name: delivery
            cardinality: many
            locate:
              tab: Schedule
              orientation: flat
              header_row: 0
              skip_row_if:
                - column: label
                  value_pattern: "^(MONDAY|TUESDAY|WEDNESDAY|THURSDAY|FRIDAY|SATURDAY|SUNDAY)$"
            fields:
              - { name: label, source_column: label, type: string }
              - { name: qty, source_column: qty, type: integer }
        """,
        tmp_path,
    )

    assert len(result.canonical["deliveries"]) == 3


def test_skip_row_if_drops_grand_total_with_compound_predicate(tmp_path):
    """A grand-total row has blank discriminator AND populated numeric column —
    geometrically identical to a real row except for the discriminator.
    """

    def build(wb):
        ws = wb.create_sheet("Tot")
        ws.append(["site", "head_count"])
        ws.append(["Site-A", 100])
        ws.append(["Site-B", 200])
        ws.append(["Site-C", 300])
        ws.append([None, 600])  # grand total: blank site, sum of head_count

    xlsx = _xlsx(tmp_path, build)
    result = _run(
        xlsx,
        """
        template_id: skip_row_if_grand_total
        version: 1
        description: P0-2 fixture - drop grand-total compound predicate
        entities:
          - name: site_count
            cardinality: many
            locate:
              tab: Tot
              orientation: flat
              header_row: 0
              skip_row_if:
                - all_blank: [site]
                  non_blank: [head_count]
            fields:
              - { name: site, source_column: site, type: string }
              - { name: head_count, source_column: head_count, type: integer }
        """,
        tmp_path,
    )

    assert len(result.canonical["site_counts"]) == 3
    assert sum(r["head_count"] for r in result.canonical["site_counts"]) == 600


# ======================================================================
# P0-3  header_levels: multi-row / merged header support (CONVENTIONS §3)
# ======================================================================


@pytest.mark.xfail(
    strict=True,
    reason="P0-3: header_levels (multi-row header) not yet implemented (CONVENTIONS §3, v2).",
)
def test_multi_row_header_combines_two_rows_into_semantic_name(tmp_path):
    """``DRY OR`` on row 0 + ``LIQUID`` on row 1 should bind as a single
    field whose source label is ``DRY OR LIQUID``.
    """

    def build(wb):
        ws = wb.create_sheet("Sheet1")
        # Two-row header.
        ws.append(["Farm", None, "DRY OR"])
        ws.append([None, "House", "LIQUID"])
        ws.append(["Foo Farms", "01", "LIQUID"])
        ws.append(["Bar Acres", "02", "DRY"])

    xlsx = _xlsx(tmp_path, build)
    result = _run(
        xlsx,
        """
        template_id: header_levels_basic
        version: 1
        entities:
          - name: row
            cardinality: many
            locate:
              tab: Sheet1
              orientation: flat
              header_levels: 2
              header_row: 0
            fields:
              - { name: farm, source_column: "Farm", type: string }
              - { name: house, source_column: "House", type: string }
              - { name: dry_or_liquid, source_column: "DRY OR LIQUID", type: string }
        """,
        tmp_path,
    )

    assert [r["dry_or_liquid"] for r in result.canonical["rows"]] == ["LIQUID", "DRY"]


def test_header_above_nonblank_emits_ambiguous_warning(tmp_path):
    """If ``header_row: 1`` and row 0 is non-empty with text that would change
    the semantic name (e.g. ``EST.`` above ``NUMBER``), surface a structured
    ``header_above_nonblank`` warning rather than silently extracting against
    only the bottom row.
    """

    def build(wb):
        ws = wb.create_sheet("Sheet1")
        ws.append([None, "EST."])  # row above header carries qualifying text
        ws.append(["farm", "NUMBER"])
        ws.append(["Foo Farms", 100])

    xlsx = _xlsx(tmp_path, build)
    tmpl = _load(
        tmp_path,
        """
        template_id: header_above_nonblank
        version: 1
        description: P0-3 interim - warn when row above header_row has content
        entities:
          - name: row
            cardinality: many
            locate:
              tab: Sheet1
              orientation: flat
              header_row: 1
            fields:
              - { name: farm, source_column: "farm", type: string }
              - { name: number, source_column: "NUMBER", type: integer }
        """,
    )
    result = extract(xlsx, tmpl)
    report = validate(result, tmpl)
    assert any(e.type == "header_above_nonblank" for e in report.errors()), "expected a header_above_nonblank warning"


# ======================================================================
# P0-4  header normalization should collapse internal whitespace/newlines
# ======================================================================


def test_header_normalization_collapses_newlines(tmp_path):
    """A header cell with an Excel line-wrap (``"Total \\nEggs"``) should
    match ``source_column: "Total Eggs"`` after normalization.
    """

    def build(wb):
        ws = wb.create_sheet("Sheet1")
        ws.append(["Farm", "Total \nEggs"])  # literal newline in header
        ws.append(["Foo Farms", 42])

    xlsx = _xlsx(tmp_path, build)
    result = _run(
        xlsx,
        """
        template_id: header_normalize_newlines
        version: 1
        description: P0-4 fixture - header with embedded newline
        entities:
          - name: row
            cardinality: many
            locate:
              tab: Sheet1
              orientation: flat
              header_row: 0
            fields:
              - { name: farm, source_column: "Farm", type: string }
              - { name: total_eggs, source_column: "Total Eggs", type: integer }
        """,
        tmp_path,
    )

    assert result.canonical["rows"][0]["total_eggs"] == 42


def test_header_normalization_collapses_double_spaces(tmp_path):
    def build(wb):
        ws = wb.create_sheet("Sheet1")
        ws.append(["Farm", "House  No"])  # double-space typo
        ws.append(["Foo Farms", "03"])

    xlsx = _xlsx(tmp_path, build)
    result = _run(
        xlsx,
        """
        template_id: header_normalize_double_space
        version: 1
        description: P0-4 fixture - header with double space
        entities:
          - name: row
            cardinality: many
            locate:
              tab: Sheet1
              orientation: flat
              header_row: 0
            fields:
              - { name: farm, source_column: "Farm", type: string }
              - { name: house_no, source_column: "House No", type: string }
        """,
        tmp_path,
    )

    assert result.canonical["rows"][0]["house_no"] == "03"


# ======================================================================
# P1-1  Anchor column / column_range / nth (disambiguation)
# ======================================================================


def test_anchor_column_scopes_match_to_one_column(tmp_path):
    """Two side-by-side blocks (REPORTING in col A, BILLING in col D) carry
    the same labels. ``anchor.column: 3`` should restrict the search to the
    billing block.
    """

    def build(wb):
        ws = wb.create_sheet("Sheet1")
        ws.append(["REPORTING", None, None, "BILLING", None, None])
        ws.append(["Company:", "Acme Co.", None, "Company:", "Globex Corp", None])
        ws.append(["Email:", "a@example.com", None, "Email:", "b@example.com", None])

    xlsx = _xlsx(tmp_path, build)
    result = _run(
        xlsx,
        """
        template_id: anchor_column_scope
        version: 1
        description: P1-1 fixture - anchor.column scopes label search
        entities:
          - name: cover
            cardinality: one
            locate:
              tab: Sheet1
              orientation: anchored
            fields:
              - name: reporting_company
                type: string
                anchor: { label_match: "Company:", column: 0, value_at: right, offset: 1 }
              - name: billing_company
                type: string
                anchor: { label_match: "Company:", column: 3, value_at: right, offset: 1 }
        """,
        tmp_path,
    )

    assert result.canonical["cover"]["reporting_company"] == "Acme Co."
    assert result.canonical["cover"]["billing_company"] == "Globex Corp"


def test_anchor_nth_picks_second_match(tmp_path):
    """A label ``SHIPPING INFORMATION`` appears twice on the sheet (a header
    label at row 0 and a sub-section label at row 4). ``nth: 2`` should pick
    the second.
    """

    def build(wb):
        ws = wb.create_sheet("Sheet1")
        ws.append(["SHIPPING INFORMATION"])
        ws.append([None])
        ws.append([None])
        ws.append([None])
        ws.append(["SHIPPING INFORMATION", "Carrier", "FedEx Generic"])

    xlsx = _xlsx(tmp_path, build)
    result = _run(
        xlsx,
        """
        template_id: anchor_nth
        version: 1
        description: P1-1 fixture - anchor.nth picks second match
        entities:
          - name: cover
            cardinality: one
            locate:
              tab: Sheet1
              orientation: anchored
            fields:
              - name: carrier_after_label
                type: string
                anchor:
                  label_match: "SHIPPING INFORMATION"
                  nth: 2
                  value_at: right
                  offset: 2
        """,
        tmp_path,
    )

    assert result.canonical["cover"]["carrier_after_label"] == "FedEx Generic"


# ======================================================================
# P1-2  Duplicate header detection
# ======================================================================


def test_duplicate_source_column_warns_when_header_appears_twice(tmp_path):
    """If two header cells in the same row have the same normalized value
    (e.g. ``DATE`` in col 0 and ``DATE`` in col 5), a template that binds
    two fields to ``source_column: DATE`` should produce a structured
    ``header_duplicated`` warning rather than silently binding both to the
    first column.
    """

    def build(wb):
        ws = wb.create_sheet("Sheet1")
        ws.append(["DATE", "qty", None, None, None, "DATE"])
        ws.append([date(2026, 1, 1), 10, None, None, None, date(2026, 1, 5)])

    xlsx = _xlsx(tmp_path, build)
    tmpl = _load(
        tmp_path,
        """
        template_id: duplicate_header
        version: 1
        description: P1-2 fixture - same header text in two columns
        entities:
          - name: row
            cardinality: many
            locate:
              tab: Sheet1
              orientation: flat
              header_row: 0
            fields:
              - { name: open_date,  source_column: "DATE", type: date }
              - { name: close_date, source_column: "DATE", type: date }
              - { name: qty,        source_column: "qty",  type: integer }
        """,
    )
    result = extract(xlsx, tmpl)
    report = validate(result, tmpl)

    assert any(e.type == "header_duplicated" for e in report.errors()), "expected header_duplicated warning"


def test_source_column_index_binds_to_second_occurrence(tmp_path):
    """Explicit ``source_column_index: 1`` should bind to the *second*
    occurrence of a duplicated header (0-indexed within matches).
    """

    def build(wb):
        ws = wb.create_sheet("Sheet1")
        ws.append(["DATE", "DATE"])
        ws.append([date(2026, 1, 1), date(2026, 1, 5)])

    xlsx = _xlsx(tmp_path, build)
    result = _run(
        xlsx,
        """
        template_id: source_column_index
        version: 1
        description: P1-2 fixture - source_column_index tiebreaker
        entities:
          - name: row
            cardinality: many
            locate:
              tab: Sheet1
              orientation: flat
              header_row: 0
            fields:
              - { name: first,  source_column: "DATE", source_column_index: 0, type: date }
              - { name: second, source_column: "DATE", source_column_index: 1, type: date }
        """,
        tmp_path,
    )

    assert result.canonical["rows"][0]["first"] == "2026-01-01"
    assert result.canonical["rows"][0]["second"] == "2026-01-05"


# ======================================================================
# P1-3  No `time` field type
# ======================================================================


@pytest.mark.xfail(
    strict=True,
    reason="P1-3: 'time' is not in FieldType; datetime.time cells need their own type.",
)
def test_time_type_accepts_native_datetime_time(tmp_path):
    def build(wb):
        ws = wb.create_sheet("Sheet1")
        ws.append(["sample", "collection_time"])
        ws.append(["S-001", time(9, 30)])
        ws.append(["S-002", time(10, 30)])

    xlsx = _xlsx(tmp_path, build)
    result = _run(
        xlsx,
        """
        template_id: time_native
        version: 1
        entities:
          - name: sample
            cardinality: many
            locate:
              tab: Sheet1
              orientation: flat
              header_row: 0
            fields:
              - { name: sample, source_column: "sample", type: string }
              - { name: collection_time, source_column: "collection_time", type: time }
        """,
        tmp_path,
    )

    assert result.canonical["samples"][0]["collection_time"] == time(9, 30)
    assert result.canonical["samples"][1]["collection_time"] == time(10, 30)


@pytest.mark.xfail(
    strict=True,
    reason="P1-3b: time_format coercion for free-text time strings not implemented.",
)
def test_time_type_coerces_free_text_strings(tmp_path):
    """``"7:30 a.m."`` should coerce to ``time(7, 30)`` given a time_format."""

    def build(wb):
        ws = wb.create_sheet("Sheet1")
        ws.append(["sample", "collection_time"])
        ws.append(["S-001", "7:30 a.m."])
        ws.append(["S-002", "12:00 p.m."])

    xlsx = _xlsx(tmp_path, build)
    result = _run(
        xlsx,
        """
        template_id: time_format
        version: 1
        entities:
          - name: sample
            cardinality: many
            locate:
              tab: Sheet1
              orientation: flat
              header_row: 0
            fields:
              - { name: sample, source_column: "sample", type: string }
              - name: collection_time
                source_column: "collection_time"
                type: time
                time_formats:
                  - "%I:%M %p"
                  - "%I:%M %P"
                  - "%-I:%M %p"
        """,
        tmp_path,
    )

    assert result.canonical["samples"][0]["collection_time"] == time(7, 30)
    assert result.canonical["samples"][1]["collection_time"] == time(12, 0)


# ======================================================================
# P1-4  Anchored entity: distinguish "label missing" from "value blank"
# ======================================================================


@pytest.mark.xfail(
    strict=True,
    reason="P1-4: anchored field returns null indistinguishably for 'label missing' "
    "and 'label present, value blank'.",
)
def test_anchor_label_was_marks_label_presence(tmp_path):
    """Two files: one with the label and a blank value, one with no label
    at all. The proposed contract: ``ctx.label_was: 'present'`` in the
    first case, ``ctx.label_was: 'absent'`` in the second.
    """

    def build_label_present(wb):
        ws = wb.create_sheet("Sheet1")
        ws.append(["Organization:", None])  # label, no value

    def build_label_absent(wb):
        ws = wb.create_sheet("Sheet1")
        ws.append([None, None])  # nothing

    tmpl_yml = """
        template_id: anchor_label_was
        version: 1
        entities:
          - name: cover
            cardinality: one
            locate:
              tab: Sheet1
              orientation: anchored
            fields:
              - name: organization
                type: string
                nullable: true
                anchor: { label_match: "Organization:", value_at: right, offset: 1 }
        """

    xlsx_a = _xlsx(tmp_path, build_label_present, name="a.xlsx")
    xlsx_b = _xlsx(tmp_path, build_label_absent, name="b.xlsx")
    tmpl = _load(tmp_path, tmpl_yml)
    report_a = validate(extract(xlsx_a, tmpl), tmpl)
    report_b = validate(extract(xlsx_b, tmpl), tmpl)

    a_errs = [e for e in report_a.errors() if e.loc[-1] == "organization"]
    b_errs = [e for e in report_b.errors() if e.loc[-1] == "organization"]

    assert a_errs and a_errs[0].ctx.get("label_was") == "present"
    assert b_errs and b_errs[0].ctx.get("label_was") == "absent"


# ======================================================================
# P1-5  cell_range honored by the extractor (CONVENTIONS §5, v1.5)
# ======================================================================


def test_cell_range_restricts_extraction_to_subrectangle(tmp_path):
    """One tab carries two unrelated tables in disjoint column ranges (cols
    A-B and cols D-E). Two entities with disjoint ``cell_range`` should each
    see only their own table.
    """

    def build(wb):
        ws = wb.create_sheet("Sheet1")
        ws.append(["site", "qty", None, "lookup_site", "head"])
        ws.append(["Site-A", 10, None, "Site-X", 1000])
        ws.append(["Site-B", 20, None, "Site-Y", 2000])

    xlsx = _xlsx(tmp_path, build)
    result = _run(
        xlsx,
        """
        template_id: cell_range_side_by_side
        version: 1
        description: P1-5 fixture - side-by-side tables via cell_range
        entities:
          - name: left_table
            cardinality: many
            locate:
              tab: Sheet1
              orientation: flat
              header_row: 0
              cell_range: A1:B3
            fields:
              - { name: site, source_column: "site", type: string }
              - { name: qty, source_column: "qty", type: integer }
          - name: right_table
            cardinality: many
            locate:
              tab: Sheet1
              orientation: flat
              header_row: 0
              cell_range: D1:E3
            fields:
              - { name: site, source_column: "lookup_site", type: string }
              - { name: head, source_column: "head", type: integer }
        """,
        tmp_path,
    )

    left = result.canonical["left_tables"]
    right = result.canonical["right_tables"]
    assert [r["site"] for r in left] == ["Site-A", "Site-B"]
    assert [r["site"] for r in right] == ["Site-X", "Site-Y"]


# ======================================================================
# P2-1  null_tokens regex/pattern support
# ======================================================================


@pytest.mark.xfail(
    strict=True,
    reason="P2-1: null_patterns (regex) not yet implemented; null_tokens accepts " "exact strings only.",
)
def test_null_patterns_match_unfilled_form_placeholders(tmp_path):
    """A regex ``^\\[.+\\]$`` should collapse any bracketed placeholder
    (``[Company]``, ``[Email]``, ``[Fax]``) to ``None`` without enumerating
    each variant.
    """

    def build(wb):
        ws = wb.create_sheet("Sheet1")
        ws.append(["company", "email"])
        ws.append(["[Company]", "[Email]"])
        ws.append(["Acme Co.", "a@example.com"])

    xlsx = _xlsx(tmp_path, build)
    result = _run(
        xlsx,
        """
        template_id: null_patterns_brackets
        version: 1
        null_patterns:
          - "^\\\\[.+\\\\]$"
        entities:
          - name: row
            cardinality: many
            locate:
              tab: Sheet1
              orientation: flat
              header_row: 0
            fields:
              - { name: company, source_column: "company", type: string, nullable: true }
              - { name: email,   source_column: "email",   type: email,  nullable: true }
        """,
        tmp_path,
    )

    assert result.canonical["rows"][0] == {"company": None, "email": None}
    assert result.canonical["rows"][1] == {"company": "Acme Co.", "email": "a@example.com"}


# ======================================================================
# P2-2  forward_fill / row inherits column values from previous row
# ======================================================================


@pytest.mark.xfail(
    strict=True,
    reason="P2-2: forward_fill not yet implemented; continuation rows whose group "
    "columns are blank emit as standalone records.",
)
def test_forward_fill_inherits_group_columns_from_previous_row(tmp_path):
    """A schedule with day-of-week + grower set on the first row of a group,
    blank on continuation rows. ``forward_fill`` should propagate the values
    downward until the next non-blank.
    """

    def build(wb):
        ws = wb.create_sheet("Sheet1")
        ws.append(["day", "grower", "houses", "qty"])
        ws.append(["MONDAY", "Acme Co.", "1-4", 100])
        ws.append([None, None, "5-6", 50])  # continues Acme/MONDAY
        ws.append(["TUESDAY", "Globex Corp", "1-4", 80])
        ws.append([None, None, "5-8", 40])  # continues Globex/TUESDAY

    xlsx = _xlsx(tmp_path, build)
    result = _run(
        xlsx,
        """
        template_id: forward_fill
        version: 1
        entities:
          - name: placement
            cardinality: many
            locate:
              tab: Sheet1
              orientation: flat
              header_row: 0
              forward_fill: [day, grower]
            fields:
              - { name: day, source_column: "day", type: string }
              - { name: grower, source_column: "grower", type: string }
              - { name: houses, source_column: "houses", type: string }
              - { name: qty, source_column: "qty", type: integer }
        """,
        tmp_path,
    )

    rows = result.canonical["placements"]
    assert len(rows) == 4
    assert [r["day"] for r in rows] == ["MONDAY", "MONDAY", "TUESDAY", "TUESDAY"]
    assert [r["grower"] for r in rows] == ["Acme Co.", "Acme Co.", "Globex Corp", "Globex Corp"]


# ======================================================================
# P2-3  data_ends_at: value_pattern (regex) instead of exact-string value
# ======================================================================


@pytest.mark.xfail(
    strict=True,
    reason="P2-3: data_ends_at supports only exact-string `value`, not regex `value_pattern`.",
)
def test_data_ends_at_value_pattern_stops_on_regex_match(tmp_path):
    """A sentinel row whose col 0 reads ``AVG AGE  25+ :`` (double-space, trailing
    space-colon — operator typo) should stop the read via a regex on
    ``data_ends_at``.
    """

    def build(wb):
        ws = wb.create_sheet("Sheet1")
        ws.append(["label", "value"])
        ws.append(["Foo Farms", 100])
        ws.append(["Bar Acres", 200])
        ws.append(["AVG AGE  25+ :", 150])  # sentinel — stop here
        ws.append(["Age 25+ Total:", 300])

    xlsx = _xlsx(tmp_path, build)
    result = _run(
        xlsx,
        """
        template_id: data_ends_at_regex
        version: 1
        entities:
          - name: row
            cardinality: many
            locate:
              tab: Sheet1
              orientation: flat
              header_row: 0
              data_ends_at:
                type: value_pattern
                column: 0
                value_pattern: "^AVG AGE\\\\s+\\\\d+\\\\+\\\\s*:?\\\\s*$"
            fields:
              - { name: label, source_column: "label", type: string }
              - { name: value, source_column: "value", type: integer }
        """,
        tmp_path,
    )

    assert [r["label"] for r in result.canonical["rows"]] == ["Foo Farms", "Bar Acres"]


# ======================================================================
# P2-4  Per-row enrichment from anchored cells on the same tab
# ======================================================================


@pytest.mark.xfail(
    strict=True,
    reason="P2-4: per-row enrich from anchored cells not yet implemented; "
    "anchor is field-level on cardinality:one entities only.",
)
def test_enrich_from_anchor_attaches_label_value_to_every_row(tmp_path):
    """A header block above the data has ``PROJECTED HATCH: 2026-01-15`` —
    the date applies to every detail row below. Should bind as an
    ``enrich`` field with ``source: anchor``.
    """

    def build(wb):
        ws = wb.create_sheet("Sheet1")
        ws.append(["PROJECTED HATCH:", date(2026, 1, 15), None])
        ws.append([None, None, None])
        ws.append(["farm", "house", "head"])
        ws.append(["Foo Farms", "01", 1000])
        ws.append(["Foo Farms", "02", 1200])
        ws.append(["Bar Acres", "01", 800])

    xlsx = _xlsx(tmp_path, build)
    result = _run(
        xlsx,
        """
        template_id: enrich_from_anchor
        version: 1
        entities:
          - name: placement
            cardinality: many
            locate:
              tab: Sheet1
              orientation: flat
              header_row: 2
            enrich:
              - field: projected_hatch
                source: anchor
                label_match: "PROJECTED HATCH:"
                value_at: right
                offset: 1
                type: date
            fields:
              - { name: farm, source_column: "farm", type: string }
              - { name: house, source_column: "house", type: string }
              - { name: head, source_column: "head", type: integer }
        """,
        tmp_path,
    )

    for row in result.canonical["placements"]:
        assert row["projected_hatch"] == date(2026, 1, 15)


# ======================================================================
# P2-5  File-open failures as Report.errors() entries (not typed exceptions)
# ======================================================================


@pytest.mark.xfail(
    strict=True,
    reason="P2-5: corrupted source files raise SourceFileError instead of surfacing as "
    "a Report with a structural `unreadable_source` error.",
)
def test_corrupted_xlsx_surfaces_as_unreadable_source_error(tmp_path):
    """A truncated zip should produce ``Report.errors()`` containing a
    structural ``unreadable_source`` entry — uniform with other structural
    failures.
    """
    # Build a deliberately-invalid xlsx (truncated zip — no EOCD).
    bad = tmp_path / "bad.xlsx"
    bad.write_bytes(b"PK\x03\x04" + b"\x00" * 16)  # local file header only, no central dir
    assert not zipfile.is_zipfile(bad)

    tmpl = _load(
        tmp_path,
        """
        template_id: unreadable_source
        version: 1
        entities:
          - name: row
            cardinality: many
            locate:
              tab: Sheet1
              orientation: flat
              header_row: 0
            fields:
              - { name: x, source_column: "x", type: string }
        """,
    )

    # Proposed contract: extract() never raises for file-format issues; it
    # returns a Result whose Report carries a structural error.
    result = extract(bad, tmpl)
    report = validate(result, tmpl)
    assert any(e.type == "unreadable_source" for e in report.errors())
    assert report.has_structural


# ======================================================================
# P2-6  `crease classify` — judge "does this file resemble this template?"
# ======================================================================


@pytest.mark.xfail(
    strict=True,
    reason="P2-6: crease.classify() does not exist; needed for a 'this file shape "
    "isn't templatable' verdict before attempting extraction.",
)
def test_classify_reports_low_confidence_on_unfit_file(tmp_path):
    """A file with no header row at all should classify as ``not_templatable``
    against a flat template expecting a clear header.
    """

    def build(wb):
        ws = wb.create_sheet("Sheet1")
        # No header row; pseudo-paper-form layout.
        ws.append(["MONDAY"])
        ws.append([date(2026, 1, 5), "Acme Co.", 100])
        ws.append([None, "Globex Corp", 200])
        ws.append(["TUESDAY"])

    xlsx = _xlsx(tmp_path, build)
    tmpl = _load(
        tmp_path,
        """
        template_id: classify_unfit
        version: 1
        entities:
          - name: row
            cardinality: many
            locate:
              tab: Sheet1
              orientation: flat
              header_row: 0
            fields:
              - { name: customer, source_column: "customer", type: string }
              - { name: qty, source_column: "qty", type: integer }
        """,
    )

    verdict = crease.classify(xlsx, tmpl)
    assert verdict.fit == "not_templatable"
    assert verdict.confidence < 0.3


# ======================================================================
# P2-7  Backend indexing inconsistency (skip_empty_area)
# ======================================================================


@pytest.mark.xfail(
    strict=True,
    reason="P2-7: header_row indexing diverges from python_calamine's default "
    "to_python() view when leading blank rows are present.",
)
def test_header_row_indexing_matches_repl_view_with_leading_blanks(tmp_path):
    """A workbook whose first row is blank: ``header_row: 1`` (the row index
    a user sees in the python_calamine REPL `to_python()` default view)
    should be the row that actually carries the headers.
    """

    def build(wb):
        ws = wb.create_sheet("Sheet1")
        ws.append([None, None])
        ws.append(["farm", "qty"])
        ws.append(["Foo Farms", 100])

    xlsx = _xlsx(tmp_path, build)
    result = _run(
        xlsx,
        """
        template_id: leading_blank_indexing
        version: 1
        entities:
          - name: row
            cardinality: many
            locate:
              tab: Sheet1
              orientation: flat
              header_row: 1
            fields:
              - { name: farm, source_column: "farm", type: string }
              - { name: qty, source_column: "qty", type: integer }
        """,
        tmp_path,
    )

    assert result.canonical["rows"] == [{"farm": "Foo Farms", "qty": 100}]


# ======================================================================
# P2-8  duplicate_policy per entity (legitimate cross-tab repetition)
# ======================================================================


@pytest.mark.xfail(
    strict=True,
    reason="P2-8: duplicate_policy not yet implemented; cross-tab repetition that "
    "is intentional always emits duplicate_row errors.",
)
def test_duplicate_policy_ignore_suppresses_intentional_repeats(tmp_path):
    """When a workbook intentionally repeats the same record across tabs
    (e.g. a rolling 6-week window where each week's tab carries the same
    farm-house placement), ``duplicate_policy: ignore`` should suppress
    ``duplicate_row`` errors.
    """

    def build(wb):
        for tab in ("W1", "W2"):
            ws = wb.create_sheet(tab)
            ws.append(["farm", "house", "head"])
            ws.append(["Foo Farms", "01", 1000])  # identical across both tabs

    xlsx = _xlsx(tmp_path, build)
    tmpl = _load(
        tmp_path,
        """
        template_id: duplicate_policy
        version: 1
        entities:
          - name: placement
            cardinality: many
            locate:
              tab_pattern: "^W\\\\d+$"
              orientation: flat
              header_row: 0
              duplicate_policy: ignore
            fields:
              - { name: farm, source_column: "farm", type: string }
              - { name: house, source_column: "house", type: string }
              - { name: head, source_column: "head", type: integer }
        """,
    )
    result = extract(xlsx, tmpl)
    report = validate(result, tmpl)
    assert not any(e.type == "duplicate_row" for e in report.errors())
    assert len(result.canonical["placements"]) == 2


# ======================================================================
# P2-9  Free-text annotation rows (single-cell rows)
# ======================================================================


@pytest.mark.xfail(
    strict=True,
    reason="P2-9: row_is_annotation_if (single-column-populated heuristic) "
    "not yet implemented; free-text rows emit as records with one populated field.",
)
def test_row_is_annotation_drops_single_cell_rows(tmp_path):
    """An ``-- REVISED --`` banner row in col A with everything else blank
    should be dropped, not emitted as a record with most fields null.
    """

    def build(wb):
        ws = wb.create_sheet("Sheet1")
        ws.append(["farm", "house", "head"])
        ws.append(["-- REVISED --", None, None])  # annotation banner
        ws.append(["Foo Farms", "01", 1000])

    xlsx = _xlsx(tmp_path, build)
    result = _run(
        xlsx,
        """
        template_id: row_is_annotation_if
        version: 1
        entities:
          - name: placement
            cardinality: many
            locate:
              tab: Sheet1
              orientation: flat
              header_row: 0
              row_is_annotation_if:
                - only_columns_populated: 1
            fields:
              - { name: farm, source_column: "farm", type: string }
              - { name: house, source_column: "house", type: string }
              - { name: head, source_column: "head", type: integer }
        """,
        tmp_path,
    )

    rows = result.canonical["placements"]
    assert len(rows) == 1
    assert rows[0]["farm"] == "Foo Farms"


# ======================================================================
# P2-10  tab: only / tab_index shorthand for single-data-tab workbooks
# ======================================================================


@pytest.mark.xfail(
    strict=True,
    reason="P2-10: tab: only / tab_index not yet implemented; dynamic-name single-tab "
    "workbooks need brittle tab_pattern: '.*' workarounds.",
)
def test_tab_only_binds_to_single_data_tab_regardless_of_name(tmp_path):
    def build(wb):
        # One tab, but its name varies per file. Operator named it the date.
        ws = wb.create_sheet("4-20-26")
        ws.append(["farm", "head"])
        ws.append(["Foo Farms", 1000])

    xlsx = _xlsx(tmp_path, build)
    result = _run(
        xlsx,
        """
        template_id: tab_only
        version: 1
        entities:
          - name: row
            cardinality: many
            locate:
              tab: only
              orientation: flat
              header_row: 0
            fields:
              - { name: farm, source_column: "farm", type: string }
              - { name: head, source_column: "head", type: integer }
        """,
        tmp_path,
    )

    assert result.canonical["rows"][0]["farm"] == "Foo Farms"


# ======================================================================
# P2-11  min_data_density warning (structural-noise detection)
# ======================================================================


@pytest.mark.xfail(
    strict=True,
    reason="P2-11: min_data_density warning not yet implemented; when 'extracted N "
    "records but most are mostly-blank', emit a structured suspicion.",
)
def test_min_data_density_warns_when_most_rows_mostly_blank(tmp_path):
    """If the extracted entity has rows where on average <30% of fields
    are populated, emit a ``low_data_density`` warning.
    """

    def build(wb):
        ws = wb.create_sheet("Sheet1")
        ws.append(["a", "b", "c", "d"])
        ws.append(["Foo Farms", 1, 2, 3])  # 4/4 populated
        ws.append([None, None, None, 10])  # 1/4
        ws.append([None, None, None, 20])  # 1/4
        ws.append([None, None, None, 30])  # 1/4
        ws.append([None, None, None, 40])  # 1/4

    xlsx = _xlsx(tmp_path, build)
    tmpl = _load(
        tmp_path,
        """
        template_id: min_data_density
        version: 1
        entities:
          - name: row
            cardinality: many
            locate:
              tab: Sheet1
              orientation: flat
              header_row: 0
              min_data_density: 0.5
            fields:
              - { name: a, source_column: "a", type: string, nullable: true }
              - { name: b, source_column: "b", type: integer, nullable: true }
              - { name: c, source_column: "c", type: integer, nullable: true }
              - { name: d, source_column: "d", type: integer, nullable: true }
        """,
    )
    result = extract(xlsx, tmpl)
    report = validate(result, tmpl)
    assert any(e.type == "low_data_density" for e in report.errors())


# ======================================================================
# P2-12  likely_cause: excel_time_only_cell on wrong_type for time vs datetime
# ======================================================================


@pytest.mark.xfail(
    strict=True,
    reason="P2-12: wrong_type on (datetime.time vs datetime field) has no "
    "likely_cause; should be excel_time_only_cell (symmetric to excel_autoconvert).",
)
def test_wrong_type_time_vs_datetime_emits_likely_cause(tmp_path):
    def build(wb):
        ws = wb.create_sheet("Sheet1")
        ws.append(["sample", "ts"])
        ws.append(["S-001", time(9, 30)])  # bare time-of-day, not datetime

    xlsx = _xlsx(tmp_path, build)
    tmpl = _load(
        tmp_path,
        """
        template_id: time_vs_datetime_likely_cause
        version: 1
        entities:
          - name: sample
            cardinality: many
            locate:
              tab: Sheet1
              orientation: flat
              header_row: 0
            fields:
              - { name: sample, source_column: "sample", type: string }
              - { name: ts, source_column: "ts", type: datetime }
        """,
    )
    result = extract(xlsx, tmpl)
    report = validate(result, tmpl)
    matches = [e for e in report.errors() if e.type == "wrong_type" and e.loc[-1] == "ts"]
    assert matches, "expected a wrong_type error on ts"
    assert matches[0].ctx.get("likely_cause") == "excel_time_only_cell"


# ======================================================================
# P2-13  Anchor.value_type — constraint on the matched value
# ======================================================================


@pytest.mark.xfail(
    strict=True,
    reason="P2-13: anchor.value_type not yet implemented; an anchor whose value is "
    "the wrong shape (e.g. an email where an integer was expected) still matches "
    "without protest.",
)
def test_anchor_value_type_rejects_neighbor_of_wrong_shape(tmp_path):
    """An anchor for ``Project ID:`` expects an integer neighbor, but in this
    file the operator put the project name to the right of the label.
    With ``value_type: integer`` the anchor should reject the match (or emit
    a structured ``anchor_value_type_mismatch``).
    """

    def build(wb):
        ws = wb.create_sheet("Sheet1")
        ws.append(["Project ID:", "Quarterly Sample Plan"])  # wrong neighbor (string, not int)

    xlsx = _xlsx(tmp_path, build)
    tmpl = _load(
        tmp_path,
        """
        template_id: anchor_value_type
        version: 1
        entities:
          - name: cover
            cardinality: one
            locate:
              tab: Sheet1
              orientation: anchored
            fields:
              - name: project_id
                type: integer
                anchor:
                  label_match: "Project ID:"
                  value_at: right
                  offset: 1
                  value_type: integer
        """,
    )
    result = extract(xlsx, tmpl)
    report = validate(result, tmpl)
    assert any(
        e.type == "anchor_value_type_mismatch" for e in report.errors()
    ), "expected an anchor_value_type_mismatch error"


# ======================================================================
# P2-14  `crease inspect <file>` — diagnostic of normalized header → column map
# ======================================================================


@pytest.mark.xfail(
    strict=True,
    reason="P2-14: crease.inspect_headers() helper not yet implemented; template "
    "authors need a way to see what header_row N actually contains after normalization.",
)
def test_inspect_headers_returns_normalized_header_to_index_map(tmp_path):
    """A diagnostic helper: ``crease.inspect_headers(file, tab, header_row)``
    returns the normalized header strings mapped to their column indices, so
    template authors can debug ``header_mapping_failed`` without launching
    a separate REPL.
    """

    def build(wb):
        ws = wb.create_sheet("Sheet1")
        ws.append(["Farm  Name", "Total \nEggs", "Hatch \nDate"])
        ws.append(["Foo Farms", 100, date(2026, 1, 15)])

    xlsx = _xlsx(tmp_path, build)
    headers = crease.inspect_headers(xlsx, tab="Sheet1", header_row=0)
    # After P0-4 lands too, expect collapsed whitespace.
    assert headers == {"farm name": 0, "total eggs": 1, "hatch date": 2}
