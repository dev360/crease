"""Read-backend adapter.

Two concrete backends sit behind a thin uniform API:

- `CalamineBackend` — wraps `python-calamine`. Reads `.xls`, `.xlsx`,
  `.xlsb`, and `.ods`; fast and GIL-releasing. Default for almost every
  template.
- `OpenpyxlBackend` — wraps `openpyxl`. `.xlsx` only, slower, but exposes
  cell metadata calamine does not — specifically `row_dimensions[i].hidden`,
  which the `locate.skip_hidden_rows` template feature needs.

The extractor talks to `Workbook` / `Sheet` through this module; it does
not import openpyxl or python-calamine directly.
"""

from __future__ import annotations

import warnings
from collections.abc import Iterator
from pathlib import Path
from typing import Any, Literal, Protocol

Row = list[Any]

Engine = Literal["calamine", "openpyxl"]

_OPENPYXL_SUPPORTED_SUFFIXES = frozenset({".xlsx", ".xlsm"})


class SourceFileError(Exception):
    """Raised when a workbook can't be opened by the chosen backend.

    Wraps the underlying backend exception so callers see a single
    crease-typed error type with the path and engine that produced it.
    The original exception is preserved via ``__cause__``.
    """


class Sheet(Protocol):
    """A worksheet abstracted across the two backends."""

    name: str

    def iter_rows(
        self,
        *,
        min_row: int | None = None,
        max_row: int | None = None,
        min_col: int | None = None,
        max_col: int | None = None,
    ) -> Iterator[Row]:
        """Yield value rows. Row/column bounds are 1-indexed, inclusive, matching openpyxl."""
        ...

    def hidden_row_indices(self) -> set[int]:
        """0-indexed hidden row indices. Empty set if the backend can't introspect this."""
        ...


class Workbook(Protocol):
    """A workbook abstracted across the two backends."""

    sheets: list[Sheet]

    def close(self) -> None: ...


# ---- calamine backend ---------------------------------------------------


def _normalize_cell(v: Any) -> Any:
    """Calamine returns ``""`` for empty cells and ``float`` for any number.

    openpyxl returns ``None`` for empty cells and ``int`` for numbers that
    Excel stored without a fractional part. Normalize so the extractor sees
    the same shape from either backend.
    """
    if isinstance(v, str) and not v:
        return None
    if isinstance(v, float) and not isinstance(v, bool):
        if v.is_integer():
            return int(v)
    return v


class _CalamineSheet:
    def __init__(self, calamine_sheet: Any) -> None:
        self._sheet = calamine_sheet
        self.name: str = calamine_sheet.name
        self._grid: list[Row] | None = None

    def _materialize(self) -> list[Row]:
        if self._grid is None:
            raw = self._sheet.to_python(skip_empty_area=False)
            self._grid = [[_normalize_cell(v) for v in row] for row in raw]
        return self._grid

    def iter_rows(
        self,
        *,
        min_row: int | None = None,
        max_row: int | None = None,
        min_col: int | None = None,
        max_col: int | None = None,
    ) -> Iterator[Row]:
        grid = self._materialize()
        lo_r = 0 if min_row is None else max(0, min_row - 1)
        hi_r = len(grid) if max_row is None else max_row
        lo_c = 0 if min_col is None else max(0, min_col - 1)
        hi_c = None if max_col is None else max_col

        for row in grid[lo_r:hi_r]:
            if hi_c is None:
                sliced = row[lo_c:]
            else:
                sliced = row[lo_c:hi_c]
            yield sliced

    def hidden_row_indices(self) -> set[int]:
        return set()


class _CalamineWorkbook:
    def __init__(self, path: Path) -> None:
        from python_calamine import CalamineWorkbook as _CW

        self._wb = _CW.from_path(str(path))
        names = self._visible_worksheet_names()
        self.sheets: list[Sheet] = [_CalamineSheet(self._wb.get_sheet_by_name(name)) for name in names]

    def _visible_worksheet_names(self) -> list[str]:
        """Restrict to data-bearing sheets so callers don't trip over chart/macro entries.

        Calamine surfaces every sheet in ``sheet_names`` regardless of type;
        openpyxl exposes only worksheets via ``.worksheets``. Filter the
        calamine view to match.
        """
        meta = getattr(self._wb, "sheets_metadata", None)
        if not meta:
            return list(self._wb.sheet_names)
        try:
            from python_calamine import SheetTypeEnum
        except ImportError:
            return list(self._wb.sheet_names)
        return [m.name for m in meta if getattr(m, "typ", None) == SheetTypeEnum.WorkSheet]

    def close(self) -> None:
        # python-calamine has no explicit close; let GC handle the rust handle.
        self._wb = None  # type: ignore[assignment]


# ---- openpyxl backend ---------------------------------------------------


class _OpenpyxlSheet:
    def __init__(self, ws: Any) -> None:
        self._ws = ws
        self.name: str = ws.title

    def iter_rows(
        self,
        *,
        min_row: int | None = None,
        max_row: int | None = None,
        min_col: int | None = None,
        max_col: int | None = None,
    ) -> Iterator[Row]:
        kwargs: dict[str, Any] = {"values_only": True}
        if min_row is not None:
            kwargs["min_row"] = min_row
        if max_row is not None:
            kwargs["max_row"] = max_row
        if min_col is not None:
            kwargs["min_col"] = min_col
        if max_col is not None:
            kwargs["max_col"] = max_col
        for row in self._ws.iter_rows(**kwargs):
            yield list(row)

    def hidden_row_indices(self) -> set[int]:
        hidden: set[int] = set()
        for r_idx, dim in self._ws.row_dimensions.items():
            if dim.hidden:
                hidden.add(r_idx - 1)
        return hidden


class _OpenpyxlWorkbook:
    def __init__(self, path: Path) -> None:
        import openpyxl

        self._wb = openpyxl.load_workbook(path, data_only=True, read_only=False)
        self.sheets: list[Sheet] = [_OpenpyxlSheet(ws) for ws in self._wb.worksheets]

    def close(self) -> None:
        try:
            self._wb.close()
        except Exception:
            # Best-effort cleanup: never let close mask an in-flight extraction error.
            pass


# ---- dispatch -----------------------------------------------------------


def select_engine(template: Any | None, engine: Engine | str | None) -> Engine:
    """Decide which backend to use.

    Precedence:
        1. Explicit ``engine`` kwarg wins.
        2. Templates with any entity declaring ``locate.skip_hidden_rows:
           true`` need openpyxl — only it exposes the row-hidden flag.
        3. Otherwise calamine (broader format support, faster).

    When the caller forces ``engine="calamine"`` on a template that uses
    ``skip_hidden_rows``, the feature silently degrades (calamine cannot
    introspect row visibility). A ``UserWarning`` is emitted so the
    mismatch is visible at runtime, matching crease's "fail loudly with
    coordinates" principle.
    """
    if engine is not None:
        if engine not in ("calamine", "openpyxl"):
            raise ValueError(f"engine must be 'calamine' or 'openpyxl', got {engine!r}")
        if engine == "calamine" and template is not None and _template_needs_openpyxl(template):
            warnings.warn(
                "engine='calamine' was forced for a template that uses "
                "locate.skip_hidden_rows; calamine cannot detect hidden rows, "
                "so the feature is silently a no-op. Pass engine='openpyxl' "
                "to keep skip_hidden_rows active.",
                UserWarning,
                stacklevel=2,
            )
        return engine
    if template is not None and _template_needs_openpyxl(template):
        return "openpyxl"
    return "calamine"


def _template_needs_openpyxl(template: Any) -> bool:
    for entity in getattr(template, "entities", []):
        locate = getattr(entity, "locate", None)
        if locate is not None and getattr(locate, "skip_hidden_rows", False):
            return True
    return False


def open_workbook(path: Path, engine: Engine) -> Workbook:
    """Open a workbook using the chosen backend.

    Raises:
        SourceFileError: If the file is missing, corrupt, encrypted, or in
            a format the chosen backend cannot read (e.g. ``.xls`` with
            ``engine="openpyxl"``). The original backend exception is
            preserved as ``__cause__``.
    """
    suffix = path.suffix.lower()
    if engine == "openpyxl" and suffix and suffix not in _OPENPYXL_SUPPORTED_SUFFIXES:
        raise SourceFileError(
            f"openpyxl can only read .xlsx/.xlsm; got {path.name!r} (suffix {suffix!r}). "
            "Pass engine='calamine' or convert the file to .xlsx."
        )
    if engine not in ("calamine", "openpyxl"):
        raise ValueError(f"unknown engine {engine!r}")
    try:
        if engine == "calamine":
            return _CalamineWorkbook(path)
        return _OpenpyxlWorkbook(path)
    except SourceFileError:
        raise
    except FileNotFoundError as e:
        raise SourceFileError(f"source file not found: {path}") from e
    except Exception as e:
        raise SourceFileError(
            f"could not open {path} with engine={engine!r}: {type(e).__name__}: {e}"
        ) from e
